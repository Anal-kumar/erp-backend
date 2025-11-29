from datetime import date, datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from app.db.session import get_db
from app.core.security import get_current_user
import app.models as models
import app.schemas.labour as schemas
import app.modules.users.models as user_models

router = APIRouter()

# Labour Gangs
# ============================================================

@router.post("/create_labour_gang", response_model=schemas.LabourGang)
def create_labour_gang(
    labour_gang: schemas.LabourGangCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)
):
    try:
        new_labour_gang = models.LabourGang(**labour_gang.model_dump())
        db.add(new_labour_gang)
        db.commit()
        db.refresh(new_labour_gang)
        return new_labour_gang
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/get_labour_gang", response_model=List[schemas.LabourGang])
def read_labour_gang(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    try:
        labour_gang = db.query(models.LabourGang).all()
        return labour_gang
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.put("/update_labour_gang/{gang_id}", response_model=schemas.LabourGang)
def update_labour_gang(
    gang_id: int, labour_gang: schemas.LabourGangUpdate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)
):
    try:
        db_gang = db.query(models.LabourGang).filter_by(id=gang_id).first()
        if not db_gang:
            raise HTTPException(status_code=404, detail="Labour gang not found")
        for key, value in labour_gang.model_dump().items():
            setattr(db_gang, key, value)
        db.commit()
        db.refresh(db_gang)
        return db_gang
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ============================================================
# Labour Work Items
# ============================================================

@router.post("/create_labour_work_item", response_model=schemas.LabourWorkItem)
def create_labour_work_item(
    labour_work_item: schemas.LabourWorkItemCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)
):
    try:
        new_labour_work_item = models.LabourWorkItem(**labour_work_item.model_dump())
        db.add(new_labour_work_item)
        db.commit()
        db.refresh(new_labour_work_item)
        return new_labour_work_item
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/get_labour_work_item", response_model=List[schemas.LabourWorkItem])
def read_labour_work_item(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    try:
        labour_work_item = db.query(models.LabourWorkItem).all()
        return labour_work_item
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.put("/update_labour_work_item/{work_item_id}", response_model=schemas.LabourWorkItem)
def update_labour_work_item(
    work_item_id: int, labour_work_item: schemas.LabourWorkItemUpdate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)
):
    try:
        db_work_item = db.query(models.LabourWorkItem).filter_by(id=work_item_id).first()
        if not db_work_item:
            raise HTTPException(status_code=404, detail="Labour work item not found")
        for key, value in labour_work_item.model_dump().items():
            setattr(db_work_item, key, value)
        db.commit()
        db.refresh(db_work_item)
        return db_work_item
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ============================================================
# Labour Work Particulars
# ============================================================

@router.post("/create_labour_work_particulars", response_model=schemas.LabourWorkParticulars)
def create_labour_work_particulars(
    labour_work_particulars: schemas.LabourWorkParticularsCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)
):
    try:
        new_labour_work_particulars = models.LabourWorkParticulars(**labour_work_particulars.model_dump())
        db.add(new_labour_work_particulars)
        db.commit()
        db.refresh(new_labour_work_particulars)
        return new_labour_work_particulars
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/get_labour_work_particulars_details", response_model=List[schemas.LabourWorkParticulars])
def read_labour_work_particulars(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    try:
        labour_work_particulars = db.query(models.LabourWorkParticulars).all()
        return labour_work_particulars
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.put("/update_labour_work_particulars/{work_particulars_id}", response_model=schemas.LabourWorkParticulars)
def update_labour_work_particulars(
    work_particulars_id: int, labour_work_particulars: schemas.LabourWorkParticularsUpdate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)
):
    try:
        db_work_particulars = db.query(models.LabourWorkParticulars).filter_by(id=work_particulars_id).first()
        if not db_work_particulars:
            raise HTTPException(status_code=404, detail="Labour work particulars not found")
        for key, value in labour_work_particulars.model_dump().items():
            setattr(db_work_particulars, key, value)
        db.commit()
        db.refresh(db_work_particulars)
        return db_work_particulars
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ============================================================
# Labour Work Locations
# ============================================================

@router.post("/create_labour_work_location", response_model=schemas.LabourWorkLocation)
def create_labour_work_location(
    labour_work_location: schemas.LabourWorkLocationCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)
):
    try:
        new_labour_work_location = models.LabourWorkLocation(**labour_work_location.model_dump())
        db.add(new_labour_work_location)
        db.commit()
        db.refresh(new_labour_work_location)
        return new_labour_work_location
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/get_labour_work_location_details", response_model=List[schemas.LabourWorkLocation])
def read_labour_work_location(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    try:
        labour_work_location = db.query(models.LabourWorkLocation).all()
        return labour_work_location
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.put("/update_labour_work_location/{work_location_id}", response_model=schemas.LabourWorkLocation)
def update_labour_work_location(
    work_location_id: int, labour_work_location: schemas.LabourWorkLocationUpdate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)
):
    try:
        db_work_location = db.query(models.LabourWorkLocation).filter_by(id=work_location_id).first()
        if not db_work_location:
            raise HTTPException(status_code=404, detail="Labour work location not found")
        for key, value in labour_work_location.model_dump().items():
            setattr(db_work_location, key, value)
        db.commit()
        db.refresh(db_work_location)
        return db_work_location
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ============================================================
# Labour Bag Packagings
# ============================================================

@router.post("/create_labour_bag_packaging", response_model=schemas.LabourBagPackagingWeight)
def create_labour_work_packaging(
    labour_work_packaging: schemas.LabourBagPackagingWeightCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)
):
    try:
        new_labour_work_packaging = models.LabourBagPackagingWeight(**labour_work_packaging.model_dump())
        db.add(new_labour_work_packaging)
        db.commit()
        db.refresh(new_labour_work_packaging)
        return new_labour_work_packaging
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/get_labour_bag_packaging_details", response_model=List[schemas.LabourBagPackagingWeight])
def read_labour_work_packaging(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    try:
        labour_work_packaging = db.query(models.LabourBagPackagingWeight).all()
        return labour_work_packaging
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.put("/update_labour_bag_packaging/{packaging_id}", response_model=schemas.LabourBagPackagingWeight)
def update_labour_work_packaging(
    packaging_id: int, labour_work_packaging: schemas.LabourBagPackagingWeightUpdate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)
):
    try:
        db_work_packaging = db.query(models.LabourBagPackagingWeight).filter_by(id=packaging_id).first()
        if not db_work_packaging:
            raise HTTPException(status_code=404, detail="Labour work packaging not found")
        for key, value in labour_work_packaging.model_dump().items():
            setattr(db_work_packaging, key, value)
        db.commit()
        db.refresh(db_work_packaging)
        return db_work_packaging
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ============================================================
# Labour Payment Vouchers
# ============================================================

@router.post("/create_voucher", response_model=schemas.LabourPaymentVoucher)
def create_labour_payment_voucher(
    voucher: schemas.LabourPaymentVoucherCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # Create main voucher record
    db_voucher = models.LabourPaymentVouchers(
        vch_date=voucher.vch_date,
        remarks=voucher.remarks,
        user_login_id=voucher.user_login_id,
    )
    db.add(db_voucher)
    db.commit()
    db.refresh(db_voucher)

    # Voucher Gangs
    for gang in voucher.labour_gang:
        db_gang = db.query(models.LabourGang).filter_by(
            gang_name=gang.gang_name
        ).first()
        if not db_gang:
            raise HTTPException(status_code=400, detail=f"Gang '{gang.gang_name}' does not exist.")
        db.add(models.VoucherGang(
            voucher_labour_payment_id=db_voucher.id,
            gang_id=db_gang.id,
            work_rate=gang.work_rate
        ))

    # Voucher Work Items
    for item in voucher.labour_work_item:
        db_item = db.query(models.LabourWorkItem).filter_by(
            labour_item_name=item.work_item_name
        ).first()
        if not db_item:
            raise HTTPException(status_code=400, detail=f"Work item '{item.labour_item_name}' does not exist.")
        db.add(models.VoucherWorkItem(
            voucher_labour_payment_id=db_voucher.id,
            work_item_id=db_item.id
        ))

    # Voucher Particulars
    for particular in voucher.labour_work_particulars:
        db_particular = db.query(models.LabourWorkParticulars).filter_by(
            work_name=particular.particular_name
        ).first()
        if not db_particular:
            raise HTTPException(status_code=400, detail=f"Work particular '{particular.particular_name}' does not exist.")
        db.add(models.VoucherParticular(
            voucher_labour_payment_id=db_voucher.id,
            particulars_id=db_particular.id
        ))

    # Voucher Bag Packagings
    for packaging in voucher.labour_bag_packaging_weight:
        db_packaging = db.query(models.LabourBagPackagingWeight).filter_by(
            bag_weight=packaging.bag_weight
        ).first()
        if not db_packaging:
            raise HTTPException(status_code=400, detail=f"Packaging weight '{packaging.bag_weight}' does not exist.")
        db.add(models.VoucherBagPackaging(
            voucher_labour_payment_id=db_voucher.id,
            bag_packaging_id=db_packaging.id,
            bags_nos=packaging.bags_nos,
            gang_id=db_gang.id
        ))

    # Voucher Locations (Origin/Destination pairings)
    for origin, dest in zip(voucher.labour_work_location_id_origin, voucher.labour_work_location_id_destination):
        db_origin = db.query(models.LabourWorkLocation).filter_by(
            work_locations=origin.work_locations
        ).first()
        if not db_origin:
            raise HTTPException(status_code=400, detail=f"Origin location '{origin.work_locations}' does not exist.")

        db_dest = db.query(models.LabourWorkLocation).filter_by(
            work_locations=dest.work_locations
        ).first()
        if not db_dest:
            raise HTTPException(status_code=400, detail=f"Destination location '{dest.work_locations}' does not exist.")

        db.add(models.VoucherLocation(
            voucher_labour_payment_id=db_voucher.id,
            labour_work_location_id_origin=db_origin.id,
            labour_work_location_id_destination=db_dest.id
        ))

    db.commit()
    db.refresh(db_voucher)
    return db_voucher

@router.put("/update_voucher/{voucher_labour_payment_id}", status_code=204)
def update_labour_payment_voucher(
    voucher_labour_payment_id: int,
    voucher_update: schemas.LabourPaymentVoucherUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # Fetch existing voucher
    db_voucher = db.query(models.LabourPaymentVouchers).filter_by(id=voucher_labour_payment_id).first()
    if not db_voucher:
        raise HTTPException(status_code=404, detail="Voucher not found")

    # Update scalar fields
    db_voucher.vch_date = voucher_update.vch_date
    db_voucher.remarks = voucher_update.remarks
    db_voucher.user_login_id = voucher_update.user_login_id

    # --- Helper to update nested relationships ---
    def update_nested(relationship_attr, pydantic_list, orm_model):
        existing = getattr(db_voucher, relationship_attr)
        existing.clear()  # wipe out old associations
        for item in pydantic_list:
            orm_obj = orm_model(**item.model_dump())
            existing.append(orm_obj)

    # Update associations (use new schema field names)
    update_nested("voucher_gangs", voucher_update.voucher_gangs, models.VoucherGang)
    update_nested("voucher_work_items", voucher_update.voucher_work_items, models.VoucherWorkItem)
    update_nested("voucher_particulars", voucher_update.voucher_particulars, models.VoucherParticular)
    update_nested("voucher_bag_packagings", voucher_update.voucher_bag_packagings, models.VoucherBagPackaging)
    update_nested("voucher_locations", voucher_update.voucher_locations, models.VoucherLocation)

    db.commit()
    db.refresh(db_voucher)

@router.get("/get_vouchers", response_model=List[schemas.LabourPaymentVoucher])
def get_all_labour_payment_vouchers(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    return db.query(models.LabourPaymentVouchers).options(
        joinedload(models.LabourPaymentVouchers.users).load_only(user_models.User.user_login_id),
        joinedload(models.LabourPaymentVouchers.voucher_gangs),
        joinedload(models.LabourPaymentVouchers.voucher_work_items),
        joinedload(models.LabourPaymentVouchers.voucher_particulars),
        joinedload(models.LabourPaymentVouchers.voucher_bag_packagings),
        joinedload(models.LabourPaymentVouchers.voucher_locations)
    ).all()

@router.get("/get_voucher/{voucher_labour_payment_id}", response_model=schemas.LabourPaymentVoucher)
def get_labour_payment_voucher(
    voucher_labour_payment_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    db_voucher = db.query(models.LabourPaymentVouchers).filter_by(id=voucher_labour_payment_id).first()
    if not db_voucher:
        raise HTTPException(status_code=404, detail="Voucher not found")
    return db_voucher

@router.get("/download_labour_payment_vouchers", response_class=StreamingResponse)
async def download_labour_payment_vouchers_report(
    gang_name: Optional[str] = None,
    work_item_name: Optional[str] = None,
    particular_name: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user
)):
    # Fetch all vouchers with related nested entities
    query = db.query(models.LabourPaymentVouchers).options(
        joinedload(models.LabourPaymentVouchers.users).load_only(user_models.User.user_login_id))

    if gang_name:
        query = query.join(models.LabourPaymentVouchers.voucher_gangs).join(models.VoucherGang.gang).filter(models.LabourGang.gang_name.ilike(f"%{gang_name}%"))
    if work_item_name:
        query = query.join(models.LabourPaymentVouchers.voucher_work_items).join(models.VoucherWorkItem.work_item).filter(models.LabourWorkItem.work_item_name.ilike(f"%{work_item_name}%"))
    if particular_name:
        query = query.join(models.LabourPaymentVouchers.voucher_particulars).join(models.VoucherParticular.particular).filter(models.LabourWorkParticulars.work_name.ilike(f"%{particular_name}%"))

    vouchers = query.all()

    if not vouchers:
        raise HTTPException(status_code=404, detail="No labour payment vouchers found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Labour Payment Vouchers"

    # Define headers: main fields + nested summaries
    headers = [
        "Voucher ID",
        "Voucher Date",
        "Remarks",
        "User Login ID",
        "Timestamp",
        "Gangs",
        "Work Items",
        "Particulars",
        "Bag Packagings",
        "Origin Location",
        "Destination Location"
    ]
    ws.append(headers)

    # Style headers
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Populate rows
    for v in vouchers:
        gangs = [f"{g.gang.gang_name} ({g.work_rate})" for g in v.voucher_gangs]
        work_items = [w.work_item.labour_item_name for w in v.voucher_work_items]
        particulars = [p.particular.work_name for p in v.voucher_particulars]
        bag_packagings = [f"{bp.bag_packaging.bag_weight}kg x {bp.bags_nos}" for bp in v.voucher_bag_packagings]
        origin_locations = [loc.location_origin.work_locations for loc in v.voucher_locations]
        destination_locations = [loc.location_destination.work_locations for loc in v.voucher_locations]

        max_len = max(len(gangs), len(work_items), len(particulars),
                    len(bag_packagings), len(origin_locations), len(destination_locations))

        start_row = ws.max_row + 1  # First row of this voucher

        for i in range(max_len):
            row = [
                v.id if i == 0 else "",
                v.vch_date.strftime("%Y-%m-%d") if i == 0 else "",
                v.remarks if i == 0 else "",
                getattr(v.users, "user_login_id", None) if i == 0 else "",
                v.time_stamp.strftime("%Y-%m-%d %H:%M:%S") if i == 0 else "",
                gangs[i] if i < len(gangs) else "",
                work_items[i] if i < len(work_items) else "",
                particulars[i] if i < len(particulars) else "",
                bag_packagings[i] if i < len(bag_packagings) else "",
                origin_locations[i] if i < len(origin_locations) else "",
                destination_locations[i] if i < len(destination_locations) else ""
            ]
            ws.append(row)

        end_row = ws.max_row  # Last row of this voucher

        # Merge main voucher columns vertically
        for col_idx in range(1, 6):  # Columns A-E
            ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
            cell = ws.cell(row=start_row, column=col_idx)
            cell.alignment = Alignment(vertical="center", horizontal="center")


    row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    # Then after appending rows, set number formats for date/datetime columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx, cell in enumerate(row, start=1):
            cell.fill = row_fill
            cell.border = thin_border

            if isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd HH:MM:SS"
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 5

    # Save to BytesIO
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Labour_Payment_Vouchers.xlsx"'}
    )
