from fastapi import APIRouter, status, HTTPException, Depends
from typing import Optional, List
from datetime import datetime
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
import io

from app.db.session import get_db
from app.core.security import get_current_user
import app.models as models
import app.schemas.daybook as schemas
from app.modules.users.models import User as UserModel

router = APIRouter()

@router.post("/create_daybook", response_model=schemas.DayBook)
def create_daybook(
    daybook: schemas.CreateDaybook,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    last_daybook = (
        db.query(models.DayBook)
        .order_by(models.DayBook.id.desc())
        .first()
    )
    if last_daybook is None or last_daybook.closing_bal is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Closing balance is required",
        )

    # Calculate closing balance
    if daybook.is_receipt:
        closing_bal = last_daybook.closing_bal + daybook.amount
    else:
        closing_bal = last_daybook.closing_bal - daybook.amount

    new_daybook = models.DayBook(
        opening_bal=last_daybook.closing_bal,
        closing_bal=closing_bal,
        **daybook.model_dump()
    )

    db.add(new_daybook)
    db.commit()
    db.refresh(new_daybook)
    return new_daybook

@router.get("/get_daybook", response_model=List[schemas.DayBook])
def get_daybook(
    party_name: Optional[str] = None,
    date: Optional[str] = None,
    particular: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)):
    
    query = db.query(models.DayBook).options(joinedload(models.DayBook.users))

    if party_name:
        query = query.filter(models.DayBook.party_name == party_name)

    if date:
        try:
            formatted_date = datetime.strptime(date, "%d/%m/%Y").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use dd/mm/yyyy.")
        query = query.filter(func.date(models.DayBook.ie_date) == formatted_date)

    if particular:
        query = query.filter(models.DayBook.particular == particular)

    db_daybook = query.options(joinedload(models.DayBook.users).load_only(UserModel.user_login_id)).all()
    return db_daybook

@router.get("/get_daybook/{daybook_id}", response_model=schemas.DayBook)
def get_daybook_by_id(daybook_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    db_daybook = db.query(models.DayBook).filter(models.DayBook.id == daybook_id).first()
    if db_daybook is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Daybook not found")
    return db_daybook

@router.get("/daybook_report")
def get_daybook_report_data(
    party_name: Optional[str] = None,
    particular: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    query = db.query(models.DayBook).options(joinedload(models.DayBook.users))

    if party_name:
        query = query.filter(models.DayBook.party_name.ilike(f"%{party_name}%"))

    if particular:
        query = query.filter(models.DayBook.particular.ilike(f"%{particular}%"))

    if from_date:
        try:
            from_dt = datetime.strptime(from_date, "%Y-%m-%d").date()
            query = query.filter(models.DayBook.ie_date >= from_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid from_date format. Use YYYY-MM-DD.")

    if to_date:
        try:
            to_dt = datetime.strptime(to_date, "%Y-%m-%d").date()
            query = query.filter(models.DayBook.ie_date <= to_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid to_date format. Use YYYY-MM-DD.")

    return query.all()

@router.get("/download_daybook_report")
def download_daybook_report(
    party_name: Optional[str] = None,
    particular: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    download: bool = False,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    query = db.query(models.DayBook).options(joinedload(models.DayBook.users))

    if party_name:
        query = query.filter(models.DayBook.party_name.ilike(f"%{party_name}%"))

    if particular:
        query = query.filter(models.DayBook.particular.ilike(f"%{particular}%"))

    if from_date:
        try:
            from_dt = datetime.strptime(from_date, "%Y-%m-%d").date()
            query = query.filter(models.DayBook.ie_date >= from_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid from_date format. Use YYYY-MM-DD.")

    if to_date:
        try:
            to_dt = datetime.strptime(to_date, "%Y-%m-%d").date()
            query = query.filter(models.DayBook.ie_date <= to_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid to_date format. Use YYYY-MM-DD.")

    db_daybook = query.all()

    if not download:
        return db_daybook

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daybook Report"

    headers = [
        "ID", "PARTY NAME", "PARTICULAR", "TRANSACTION DATE", "AMOUNT",
        "PAYMENT TYPE", "PAYMENT MODE", "REF NO.", "OPENING BALANCE", "CLOSING BALANCE", "ENTRY BY"
    ]
    ws.append(headers)

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for entry in db_daybook:
        is_receipt = entry.is_receipt
        payment_type = "RECEIPT" if is_receipt else "PAYMENT"
        payment_mode = "CASH" if entry.is_bank else "BANK TRANSFER"

        row_data = [
            entry.id,
            entry.party_name,
            entry.particular,
            str(entry.ie_date),
            entry.amount,
            payment_type,
            payment_mode,
            entry.ref_no,
            entry.opening_bal,
            entry.closing_bal,
            entry.users.user_login_id if entry.users else ""
        ]

        ws.append(row_data)

        current_row = ws.max_row

        # Define specific formats
        currency_format = '"₹ "#,##0.00'
        integer_format = '0'

        # Format specific cells
        ws.cell(row=current_row, column=5).number_format = currency_format  # Amount
        ws.cell(row=current_row, column=9).number_format = currency_format  # Opening Balance
        ws.cell(row=current_row, column=10).number_format = currency_format  # Closing Balance

        # Format ID column and any other integer column
        ws.cell(row=current_row, column=1).number_format = integer_format

        # Conditional color only on "PAYMENT TYPE" column (6th column)
        payment_type_cell = ws.cell(row=ws.max_row, column=6)
        if is_receipt:
            payment_type_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            payment_type_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Apply light blue background fill to table area (excluding header styling)
    light_blue_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            if cell.column != 6:  # skip PAYMENT TYPE column which already has conditional styling
                cell.fill = light_blue_fill

    # Auto column width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Add Excel Table
    table_ref = f"A1:J{ws.max_row}"
    table = Table(displayName="DaybookTable", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=daybook_report.xlsx"}
    )

@router.put("/update_daybook/{daybook_id}", response_model=schemas.DayBook)
def update_daybook(daybook_id: int, daybook: schemas.UpdateDaybook, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    db_daybook = db.query(models.DayBook).filter(models.DayBook.id == daybook_id).first()
    if db_daybook is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Daybook not found")

    if daybook_id == 1:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="First entry cannot be updated")

    prev_daybook = (
        db.query(models.DayBook)
        .filter(models.DayBook.id < daybook_id)
        .order_by(models.DayBook.id.desc())
        .first()
    )
    if prev_daybook.closing_bal is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Closing balance is required for the previous entry",
        )

    db_daybook.user_login_id = daybook.user_login_id
    db_daybook.party_name = daybook.party_name
    db_daybook.particular = daybook.particular
    db_daybook.is_bank = daybook.is_bank
    db_daybook.is_receipt = daybook.is_receipt
    db_daybook.amount = daybook.amount
    db_daybook.ref_no = daybook.ref_no
    db_daybook.remarks = daybook.remarks

    if daybook.is_bank:
        db_daybook.closing_bal = prev_daybook.closing_bal + daybook.amount
    else:
        db_daybook.closing_bal = prev_daybook.closing_bal - daybook.amount

    db.commit()
    db.refresh(db_daybook)
    return db_daybook
