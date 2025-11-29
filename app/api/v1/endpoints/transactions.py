from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload
from sqlalchemy.exc import SQLAlchemyError
from typing import List, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import io

from app.db.session import get_db
from app.core.security import get_current_user
from app.utils import calculate_financials
import app.models as models
import app.schemas.transaction as schemas
from app.modules.users.schemas import User as UserSchema
from app.schemas.transaction import BagDetailsOut, BagsStatus

router = APIRouter()

@router.get("/get_transactions", response_model=List[schemas.TransactionMillOperations])
def get_transactions(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """
    Retrieve all transaction mill operations.
    """
    transactions = db.query(models.TransactionMillOperations).options(
        joinedload(models.TransactionMillOperations.users).load_only(models.User.user_login_id),
        joinedload(models.TransactionMillOperations.bag_details).joinedload(models.BagDetails.packaging)
    ).all()
    
    return transactions

@router.get("/get_transaction/{transaction_id}", response_model=schemas.TransactionMillOperations)
def get_transaction(transaction_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """
    Retrieve a specific transaction by ID.
    """
    transaction = db.query(models.TransactionMillOperations).options(joinedload(models.TransactionMillOperations.users), joinedload(models.TransactionMillOperations.transaction_allowance_deduction_details), joinedload(models.TransactionMillOperations.transaction_payments_mill_operations)).filter(models.TransactionMillOperations.id == transaction_id).first()
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return transaction

@router.post("/create_transaction", response_model=schemas.TransactionMillOperations)
def create_transaction(
    transaction: schemas.TransactionMillOperationsCreate,
    db: Session = Depends(get_db),
    _current_user: dict = Depends(get_current_user)
):
    """
    Create a new transaction mill operation with all linked records.
    This also updates per-godown, per-item stock in StockLedger (bags + quintals).
    """
    # --- validate references (party/broker/transportor/operator) ---
    db_party = db.query(models.PartyDetails).filter_by(party_name=transaction.party_name).first()
    if not db_party:
        raise HTTPException(status_code=400, detail="The person doesn't exist in the party list.")

    db_broker = db.query(models.BrokerDetails).filter_by(broker_name=transaction.broker_name).first()
    if not db_broker:
        raise HTTPException(status_code=400, detail="The person doesn't exist in the broker list.")

    db_transportor = db.query(models.TransportorDetails).filter_by(transportor_name=transaction.transportor_name).first()
    if not db_transportor:
        raise HTTPException(status_code=400, detail="The person doesn't exist in the transportor list.")

    db_operator = db.query(models.WeightBridgeOperator).filter_by(operator_name=transaction.operator_name).first()
    if not db_operator:
        raise HTTPException(status_code=400, detail="The person doesn't exist in the weight bridge operator list.")

    # Validate Stock Items exist and map them
    stock_items = []
    for item in transaction.transaction_stock_items:
        db_item = db.query(models.StockItems).filter_by(stock_item_name=item.stock_item_name).first()
        if not db_item:
            raise HTTPException(status_code=400, detail=f"Stock item '{item.stock_item_name}' does not exist.")
        stock_items.append((db_item, item))

    # Validate Godowns exist
    for unload in transaction.unloadings:
        db_godown = db.query(models.GodownDetails).filter_by(id=unload.godown_id).first()
        if not db_godown:
            raise HTTPException(status_code=400, detail=f"Godown with ID {unload.godown_id} does not exist.")

    # --- Create main transaction record ---
    db_transaction = models.TransactionMillOperations(
        user_login_id=transaction.user_login_id,
        rst_number=transaction.rst_number,
        bill_number=transaction.bill_number,
        transaction_date=transaction.transaction_date,
        transaction_type=transaction.transaction_type,
        party_id=db_party.id,
        broker_id=db_broker.id,
        transportor_id=db_transportor.id,
        gross_weight=transaction.gross_weight,
        tare_weight=transaction.tare_weight,
        weight_bridge_operator_id=db_operator.id,
        vehicle_number=transaction.vehicle_number,
        remarks=transaction.remarks,
    )
    db.add(db_transaction)
    db.flush()  # get id without committing

    # --- Add TransactionStockItem entries ---
    # Also keep list for proportional distribution
    trans_stock_objs = []
    total_stock_bags_in_txn = 0
    for db_item, item in stock_items:
        tsi = models.TransactionStockItem(
            transaction_id=db_transaction.id,
            stock_item_id=db_item.id,
            number_of_bags=item.number_of_bags,
            weight=item.weight,
            rate=item.rate
        )
        db.add(tsi)
        trans_stock_objs.append((db_item, item))
        total_stock_bags_in_txn += (item.number_of_bags or 0)

    # --- Add Payments ---
    for payment in transaction.payments:
        db.add(models.TransactionPaymentDetails(
            transaction_id=db_transaction.id,
            payment_amount=payment.payment_amount,
            payment_date=payment.payment_date,
            payment_remarks=payment.payment_remarks
        ))

    # --- Add Packagings ---
    for packaging in transaction.packagings:
        db_packaging = db.query(models.PackagingDetails).filter_by(packaging_name=packaging.packaging_name).first()
        if not db_packaging:
            raise HTTPException(status_code=400, detail=f"Packaging '{packaging.packaging_name}' does not exist.")
        db.add(models.TransactionPackagingDetails(
            transaction_id=db_transaction.id,
            packaging_id=db_packaging.id,
            bag_nos=packaging.bag_nos
        ))

    for packaging in transaction.packagings:
        db.add(models.BagDetails(
            transaction_id=db_transaction.id,
            packaging_id=db_packaging.id,
            total_bags=packaging.bag_nos,
            returned_bags=0,
            remaining_bags=packaging.bag_nos,
            bags_status=models.BagsStatus.ACTIVE
        ))

    # --- Add Allowances/Deductions ---
    for allowance in transaction.allowances_deductions:
        db.add(models.TransactionAllowanceDeductionsDetails(
            transaction_id=db_transaction.id,
            is_allowance=allowance.is_allowance,
            allowance_deduction_name=allowance.allowance_deduction_name,
            allowance_deduction_amount=allowance.allowance_deduction_amount,
            remarks=allowance.remarks
        ))

    # --- Prepare packaging totals & per-bag net weight (kg) ---
    total_pack_bags = sum((p.bag_nos or 0) for p in transaction.packagings) if transaction.packagings else 0
    total_bag_weight_grams = 0
    for packaging in transaction.packagings:
        db_pack = db.query(models.PackagingDetails).filter_by(packaging_name=packaging.packaging_name).first()
        if db_pack:
            # db_pack.bag_weight is likely in grams per bag, so convert to kg
            total_bag_weight_grams += (db_pack.bag_weight or 0) * (packaging.bag_nos or 0)
    total_bag_weight_kg = total_bag_weight_grams / 1000.0 if total_bag_weight_grams else 0.0

    total_unload_bags = sum((u.number_of_bags or 0) for u in transaction.unloadings) if transaction.unloadings else 0
    net_weight_total_kg = (transaction.gross_weight or 0) - (transaction.tare_weight or 0)  # kg

    # 1. Weight calculation per bag (including packaging)
    weight_per_bag_including_pack_kg = (net_weight_total_kg / total_unload_bags) if total_unload_bags else 0.0
    
    # 2. Weight calculation for packaging only per bag
    bag_weight_per_bag_kg = (total_bag_weight_kg / total_pack_bags) if total_pack_bags else 0.0
    
    # 3. Weight calculation per bag (excluding packaging - i.e., net material)
    net_weight_per_bag_excluding_pack_kg = max(weight_per_bag_including_pack_kg - bag_weight_per_bag_kg, 0.0)

    # 4. Weight calculation per bag (including packaging - same as step 1, but for clarity)
    net_weight_per_bag_including_pack_kg = weight_per_bag_including_pack_kg # Use the value from step 1
    
    # --- Add Unloading Points and update godowns + stock_ledger ---
    
    for unload in transaction.unloadings:
    # get godown
        db_godown = db.query(models.GodownDetails).filter_by(id=unload.godown_id).with_for_update().first()
        if not db_godown:
            raise HTTPException(status_code=400, detail=f"Godown with ID {unload.godown_id} does not exist.")

        # Bags in this unloading
        unload_bags = (unload.number_of_bags or 0)

        # Now distribute unload_bags across stock items proportionally to their number_of_bags in transaction
        # If transaction stock bag counts don't sum (or are zero), fall back to naive equal split.
        if total_stock_bags_in_txn > 0:
            # we'll compute integer bag allocation, assign remainder to last item
            remaining_bags_to_assign = unload_bags
            for idx, (db_item, item) in enumerate(trans_stock_objs):
                if idx < len(trans_stock_objs) - 1:
                    proportion = ((item.number_of_bags or 0) / total_stock_bags_in_txn) if total_stock_bags_in_txn else 0
                    item_bags_for_unload = int(proportion * unload_bags)
                    # ensure not exceeding remaining (safety)
                    item_bags_for_unload = min(item_bags_for_unload, remaining_bags_to_assign)
                else:
                    # last item gets whatever remains
                    item_bags_for_unload = remaining_bags_to_assign

                remaining_bags_to_assign -= item_bags_for_unload

                if item_bags_for_unload <= 0:
                    continue

                # --- MODIFICATION START: Select weight based on transaction_type ---
                if transaction.transaction_type:
                    # transaction_type is TRUE (e.g., Inward/Purchase): Use NET weight (EXCLUDING packaging)
                    weight_per_bag_to_use = net_weight_per_bag_excluding_pack_kg
                else:
                    # transaction_type is FALSE (e.g., Outward/Sale): Use GROSS weight (INCLUDING packaging)
                    weight_per_bag_to_use = net_weight_per_bag_including_pack_kg
                
                # Calculate total weight in quintals for this item and unloading point
                item_weight_quintal = (weight_per_bag_to_use * item_bags_for_unload) / 100.0
                # --- MODIFICATION END ---
                
                # Update (or create) ledger entry for (godown_id, stock_item_id)
                ledger = db.query(models.StockLedger).filter_by(
                    godown_id=unload.godown_id,
                    stock_item_id=db_item.id
                ).with_for_update().first()

                if not ledger:
                    ledger = models.StockLedger(
                        godown_id=unload.godown_id,
                        stock_item_id=db_item.id,
                        stock_quantity_bags=0,
                        stock_weight_quintal=0.0
                    )
                    db.add(ledger)
                    db.flush()

                ledger.apply_stock_movement(transaction.transaction_type, item_bags_for_unload, item_weight_quintal)
        else:
            # Fallback: distribute unload bags equally among provided stock items
            if trans_stock_objs:
                
                # --- MODIFICATION START: Select weight based on transaction_type for fallback ---
                if transaction.transaction_type:
                    weight_per_bag_to_use = net_weight_per_bag_excluding_pack_kg
                else:
                    weight_per_bag_to_use = net_weight_per_bag_including_pack_kg
                # --- MODIFICATION END ---
                
                base = unload_bags // len(trans_stock_objs)
                remainder = unload_bags % len(trans_stock_objs)
                for idx, (db_item, item) in enumerate(trans_stock_objs):
                    item_bags_for_unload = base + (1 if idx < remainder else 0)
                    if item_bags_for_unload <= 0:
                        continue
                        
                    item_weight_quintal = (weight_per_bag_to_use * item_bags_for_unload) / 100.0

                    ledger = db.query(models.StockLedger).filter_by(
                        godown_id=unload.godown_id,
                        stock_item_id=db_item.id
                    ).with_for_update().first()
                    if not ledger:
                        ledger = models.StockLedger(
                            godown_id=unload.godown_id,
                            stock_item_id=db_item.id,
                            stock_quantity_bags=0,
                            stock_weight_quintal=0.0
                        )
                        db.add(ledger)
                        db.flush()

                    ledger.apply_stock_movement(transaction.transaction_type, item_bags_for_unload, item_weight_quintal)
    # --- All changes prepared; commit once ---
    db.commit()
    db.refresh(db_transaction)
    return db_transaction

@router.put("/update_transaction/{transaction_id}", response_model=schemas.TransactionMillOperations)
def update_transaction(
    transaction_id: int,
    updated_data: schemas.TransactionMillOperationsUpdate,
    db: Session = Depends(get_db),
    _current_user: dict = Depends(get_current_user)):
    """
    Update an existing mill operation transaction.
    Only reverses and reapplies stock effects when stock-related data changes.
    """

    # --- Fetch the existing transaction ---
    db_transaction = db.query(models.TransactionMillOperations).filter_by(id=transaction_id).first()
    if not db_transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")

    # --- Validate related entities ---
    db_party = db.query(models.PartyDetails).filter_by(party_name=updated_data.party_name).first()
    if not db_party:
        raise HTTPException(status_code=400, detail="Invalid party name")

    db_broker = db.query(models.BrokerDetails).filter_by(broker_name=updated_data.broker_name).first()
    if not db_broker:
        raise HTTPException(status_code=400, detail="Invalid broker name")

    db_transportor = db.query(models.TransportorDetails).filter_by(transportor_name=updated_data.transportor_name).first()
    if not db_transportor:
        raise HTTPException(status_code=400, detail="Invalid transportor name")

    db_operator = db.query(models.WeightBridgeOperator).filter_by(operator_name=updated_data.operator_name).first()
    if not db_operator:
        raise HTTPException(status_code=400, detail="Invalid operator name")

    # --- Detect if stock-affecting changes occurred ---
    old_txn_type = db_transaction.transaction_type
    old_gross = db_transaction.gross_weight or 0
    old_tare = db_transaction.tare_weight or 0

    requires_stock_update = False

    # Transaction type flip
    if updated_data.transaction_type != old_txn_type:
        requires_stock_update = True

    # Weight field change
    elif (updated_data.gross_weight or 0) != old_gross or (updated_data.tare_weight or 0) != old_tare:
        requires_stock_update = True

    # Stock item change
    elif updated_data.transaction_stock_items and (
        len(updated_data.transaction_stock_items) != len(db_transaction.transaction_stock_items)
        or any(
            (i.number_of_bags, i.weight, i.stock_item_name)
            != (db_item.number_of_bags, db_item.weight, db_item.stock_items.stock_item_name)
            for i, db_item in zip(updated_data.transaction_stock_items, db_transaction.transaction_stock_items)
        )
    ):
        requires_stock_update = True

    # Unloading change
    elif updated_data.unloadings and (
        len(updated_data.unloadings) != len(db_transaction.transaction_unloading_point_details)
        or any(
            (u.number_of_bags, u.godown_name)
            != (du.number_of_bags, du.godown.godown_name)
            for u, du in zip(updated_data.unloadings, db_transaction.transaction_unloading_point_details)
        )
    ):
        requires_stock_update = True

    # --- Update simple fields ---
    db_transaction.rst_number = updated_data.rst_number
    db_transaction.bill_number = updated_data.bill_number
    db_transaction.transaction_date = updated_data.transaction_date
    db_transaction.transaction_type = updated_data.transaction_type
    db_transaction.party_id = db_party.id
    db_transaction.broker_id = db_broker.id
    db_transaction.transportor_id = db_transportor.id
    db_transaction.weight_bridge_operator_id = db_operator.id
    db_transaction.gross_weight = updated_data.gross_weight
    db_transaction.tare_weight = updated_data.tare_weight
    db_transaction.vehicle_number = updated_data.vehicle_number
    db_transaction.remarks = updated_data.remarks

    # --- Reverse and reapply stock impact if required ---
    if requires_stock_update:
        try:
            db.begin_nested()  # Savepoint

            # Reverse old stock impact
            for unload in db_transaction.transaction_unloading_point_details:
                for stock_item in db_transaction.transaction_stock_items:
                    ledger = db.query(models.StockLedger).filter_by(
                        godown_id=unload.godown_id,
                        stock_item_id=stock_item.stock_item_id
                    ).with_for_update().first()
                    if ledger:
                        if old_txn_type:  # old purchase
                            ledger.stock_quantity_bags -= stock_item.number_of_bags
                            ledger.stock_weight_quintal -= (stock_item.weight or 0) / 100.0
                        else:  # old sale
                            ledger.stock_quantity_bags += stock_item.number_of_bags
                            ledger.stock_weight_quintal += (stock_item.weight or 0) / 100.0

            # Clear and rebuild affected children
            db_transaction.transaction_stock_items.clear()
            db_transaction.transaction_unloading_point_details.clear()
            db_transaction.transaction_packaging_details.clear()
            db.flush()

            # Stock items
            for item in updated_data.transaction_stock_items:
                db_item = db.query(models.StockItems).filter_by(stock_item_name=item.stock_item_name).first()
                if not db_item:
                    raise HTTPException(status_code=400, detail=f"Invalid stock item '{item.stock_item_name}'")
                db.add(models.TransactionStockItem(
                    transaction_id=db_transaction.id,
                    stock_item_id=db_item.id,
                    number_of_bags=item.number_of_bags,
                    weight=item.weight,
                    rate=item.rate
                ))

            # Unloadings
            for unload in updated_data.unloadings:
                godown = db.query(models.GodownDetails).filter_by(godown_name=unload.godown_name).first()
                if not godown:
                    raise HTTPException(status_code=400, detail=f"Invalid godown '{unload.godown_name}'")
                db.add(models.TransactionUnloadingPointDetails(
                    transaction_id=db_transaction.id,
                    godown_id=godown.id,
                    number_of_bags=unload.number_of_bags,
                    remarks=unload.remarks
                ))

            # Packagings
            for p in updated_data.packagings:
                db_pack = db.query(models.PackagingDetails).filter_by(packaging_name=p.packaging_name).first()
                if not db_pack:
                    raise HTTPException(status_code=400, detail=f"Invalid packaging '{p.packaging_name}'")
                db.add(models.TransactionPackagingDetails(
                    transaction_id=db_transaction.id,
                    packaging_id=db_pack.id,
                    bag_nos=p.bag_nos
                ))

            db.flush()

            # Apply new stock impact
            for unload in updated_data.unloadings:
                godown = db.query(models.GodownDetails).filter_by(godown_name=unload.godown_name).with_for_update().first()
                for item in updated_data.transaction_stock_items:
                    db_item = db.query(models.StockItems).filter_by(stock_item_name=item.stock_item_name).first()
                    ledger = db.query(models.StockLedger).filter_by(
                        godown_id=godown.id,
                        stock_item_id=db_item.id
                    ).with_for_update().first()
                    if not ledger:
                        ledger = models.StockLedger(
                            godown_id=godown.id,
                            stock_item_id=db_item.id,
                            stock_quantity_bags=0,
                            stock_weight_quintal=0.0
                        )
                        db.add(ledger)
                        db.flush()

                    if updated_data.transaction_type:  # purchase
                        ledger.stock_quantity_bags += item.number_of_bags
                        ledger.stock_weight_quintal += (item.weight or 0) / 100.0
                    else:  # sale
                        ledger.stock_quantity_bags -= item.number_of_bags
                        ledger.stock_weight_quintal -= (item.weight or 0) / 100.0

            db.commit()

        except SQLAlchemyError as e:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Stock update failed: {str(e)}")

    # --- Update non-stock child records ---
    db_transaction.transaction_allowance_deduction_details.clear()
    db_transaction.transaction_payments_mill_operations.clear()
    db.flush()

    for p in updated_data.payments:
        db.add(models.TransactionPaymentDetails(
            transaction_id=db_transaction.id,
            payment_amount=p.payment_amount,
            payment_date=p.payment_date,
            payment_remarks=p.payment_remarks
        ))

    for a in updated_data.allowances_deductions:
        db.add(models.TransactionAllowanceDeductionsDetails(
            transaction_id=db_transaction.id,
            is_allowance=a.is_allowance,
            allowance_deduction_name=a.allowance_deduction_name,
            allowance_deduction_amount=a.allowance_deduction_amount,
            remarks=a.remarks
        ))

    # --- Finalize ---
    db.commit()
    db.refresh(db_transaction)
    return db_transaction

@router.get("/stock_summary")
def get_stock_summary(
    godown_name: Optional[str] = Query(None, description="Filter by godown name"),
    stock_item_name: Optional[str] = Query(None, description="Filter by stock item name"),
    db: Session = Depends(get_db)):
    """
    Returns current stock summary (bags and quintals) per godown and per item.
    Pulls data from StockLedger and includes grand totals.

    NOTE: If stock was moved out via a sale (transaction_type == False), we subtract 
    the proportional packaging weight from the reported total weight (in quintals) 
    to show the true net material stock.
    """
    # 1. Query the raw stock ledger totals (Bags & Weight)
    query = db.query(
        models.GodownDetails.godown_name.label("godown_name"),
        models.StockItems.stock_item_name.label("stock_item_name"),
        func.sum(models.StockLedger.stock_quantity_bags).label("total_bags"),
        func.sum(models.StockLedger.stock_weight_quintal).label("total_weight_quintal")
    ).join(
        models.GodownDetails, models.GodownDetails.id == models.StockLedger.godown_id
    ).join(
        models.StockItems, models.StockItems.id == models.StockLedger.stock_item_id
    ).group_by(
        models.GodownDetails.godown_name, models.StockItems.stock_item_name
    )

    if godown_name:
        query = query.filter(models.GodownDetails.godown_name.ilike(f"%{godown_name}%"))

    if stock_item_name:
        query = query.filter(models.StockItems.stock_item_name.ilike(f"%{stock_item_name}%"))

    results = query.all()

    # 2. Compute packaging weight adjustments for all sales (transaction_type == False).
    bag_weight_adjustments = {}  # key: (godown_name, stock_item_name) -> quintals to subtract
    sales_txn_q = db.query(models.TransactionMillOperations).options(
        joinedload(models.TransactionMillOperations.transaction_packaging_details).joinedload(models.TransactionPackagingDetails.packaging),
        joinedload(models.TransactionMillOperations.transaction_unloading_point_details).joinedload(models.TransactionUnloadingPointDetails.godown),
        joinedload(models.TransactionMillOperations.transaction_stock_items).joinedload(models.TransactionStockItem.stock_items),
    ).filter(models.TransactionMillOperations.transaction_type == False)

    # Apply filters to transactions to match the summary scope
    if godown_name:
        # Note: If filtering by godown, the transaction must have at least one unloading point matching the godown filter
        sales_txn_q = sales_txn_q.join(models.TransactionUnloadingPointDetails).join(models.GodownDetails).filter(models.GodownDetails.godown_name.ilike(f"%{godown_name}%"))
    if stock_item_name:
        # Note: If filtering by stock item, the transaction must contain the stock item
        sales_txn_q = sales_txn_q.join(models.TransactionStockItem).join(models.StockItems).filter(models.StockItems.stock_item_name.ilike(f"%{stock_item_name}%"))

    sales_transactions = sales_txn_q.all()

    for tx in sales_transactions:
        pkgs = tx.transaction_packaging_details or []
        unloads = tx.transaction_unloading_point_details or []
        t_stock_items = tx.transaction_stock_items or []

        # Calculate total bag count across packagings (for bag-weight calc) and total bag-weight grams
        total_pack_bags = sum((p.bag_nos or 0) for p in pkgs) if pkgs else 0
        total_bag_weight_grams = 0
        for p in pkgs:
            if getattr(p, "packaging", None) and getattr(p.packaging, "bag_weight", None):
                total_bag_weight_grams += (p.packaging.bag_weight or 0) * (p.bag_nos or 0)

        total_unload_bags = sum((u.number_of_bags or 0) for u in unloads) if unloads else 0
        if total_unload_bags == 0 or not t_stock_items:
            continue  # nothing to distribute

        # total stock bags to determine proportional split among stock items
        total_stock_bags_in_txn = sum((si.number_of_bags or 0) for si in t_stock_items) if t_stock_items else 0
        
        # --- Distribution Logic (Handles both Proportional and Fallback Equal Split) ---
        for u in unloads:
            godown = getattr(u, "godown", None)
            if not godown:
                continue
            unload_bags = (u.number_of_bags or 0)
            
            if total_stock_bags_in_txn == 0:
                # Fallback: equal distribution of unload bags across items
                base = unload_bags // len(t_stock_items)
                remainder = unload_bags % len(t_stock_items)
                item_allocations = [(si, base + (1 if idx < remainder else 0)) for idx, si in enumerate(t_stock_items)]
            else:
                # Proportional distribution by stock item bag counts
                remaining_bags_to_assign = unload_bags
                item_allocations = []
                for idx, si in enumerate(t_stock_items):
                    if idx < len(t_stock_items) - 1:
                        proportion = ((si.number_of_bags or 0) / total_stock_bags_in_txn)
                        item_bags_for_unload = int(proportion * unload_bags)
                        item_bags_for_unload = min(item_bags_for_unload, remaining_bags_to_assign)
                    else:
                        item_bags_for_unload = remaining_bags_to_assign
                    
                    remaining_bags_to_assign -= item_bags_for_unload
                    item_allocations.append((si, item_bags_for_unload))

            # Apply allocated packaging weight
            for si, item_bags_for_unload in item_allocations:
                if item_bags_for_unload <= 0:
                    continue
                    
                # Calculate the fraction of total packaging weight to attribute to this specific godown/item chunk
                allocated_bag_weight_grams = total_bag_weight_grams * (item_bags_for_unload / total_unload_bags)
                allocated_quintal = allocated_bag_weight_grams / 100000.0 # 1000g/kg * 100kg/quintal = 100000
                
                key = (godown.godown_name, si.stock_items.stock_item_name if si.stock_items else None)
                bag_weight_adjustments[key] = bag_weight_adjustments.get(key, 0.0) + allocated_quintal

    # 3. Format per-godown totals and apply adjustment
    godown_totals = {}
    for row in results:
        godown = row.godown_name
        item_name = row.stock_item_name
        if godown not in godown_totals:
            godown_totals[godown] = {
                "godown_name": godown,
                "items": [],
                "total_bags": 0,
                "total_weight_quintal": 0.0
            }

        # Apply bag-weight adjustment if present for this godown+item (sales)
        adj = bag_weight_adjustments.get((godown, item_name), 0.0)
        adjusted_weight_quintal = float(((row.total_weight_quintal or 0) - adj))

        godown_totals[godown]["items"].append({
            "stock_item_name": item_name,
            "bags": int(row.total_bags or 0),
            "weight_quintal": adjusted_weight_quintal
        })

        godown_totals[godown]["total_bags"] += int(row.total_bags or 0)
        godown_totals[godown]["total_weight_quintal"] += adjusted_weight_quintal

    # 4. Compute grand total
    grand_total_bags = sum(g["total_bags"] for g in godown_totals.values())
    grand_total_weight = sum(g["total_weight_quintal"] for g in godown_totals.values())

    return {
        "summary": list(godown_totals.values()),
        "grand_total": {
            "total_bags": grand_total_bags,
            "total_weight_quintal": (grand_total_weight)
        }
    }

@router.post("/return_bags/{transaction_id}", response_model=List[schemas.BagDetailsOut])
def return_bags(
    transaction_id: int,
    return_data: List[schemas.BagReturnRequest],
    db: Session = Depends(get_db),
):
    updated_records = []

    for item in return_data:
        # Correct lookup: join to packaging_details and filter on real column
        bag_detail = (
            db.query(models.BagDetails)
            .join(
                models.PackagingDetails,
                models.PackagingDetails.id == models.BagDetails.packaging_id,
            )
            .filter(
                models.BagDetails.transaction_id == transaction_id,
                models.PackagingDetails.packaging_name == item.packaging_name,
            )
            .first()
        )

        if not bag_detail:
            raise HTTPException(
                status_code=404,
                detail=f"No bag record found for packaging '{item.packaging_name}' in this transaction."
            )
        
        if bag_detail.bags_status == models.BagsStatus.RETURNED:
            raise HTTPException(
                status_code=400,
                detail=f"Bags for this transaction have already been returned."
            )

        # Update returned bag count
        bag_detail.returned_bags = (bag_detail.returned_bags or 0) + item.returned_count

        # Cap it to total bags
        if bag_detail.returned_bags >= bag_detail.total_bags:
            bag_detail.returned_bags = bag_detail.total_bags
            bag_detail.bags_status = models.BagsStatus.RETURNED
        else:
            bag_detail.bags_status = models.BagsStatus.ACTIVE

        bag_detail.remaining_bags = bag_detail.total_bags - bag_detail.returned_bags

        updated_records.append(bag_detail)

    db.commit()

    for record in updated_records:
        db.refresh(record)

    return updated_records

@router.get("/download_transaction_report", response_class=StreamingResponse)
async def download_all_transactions_report(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
    party_name: Optional[str] = Query(None),
    broker_name: Optional[str] = Query(None),
    transporter_name: Optional[str] = Query(None),
    stock_item_name: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    transaction_type: Optional[bool] = Query(None),
):
    query = db.query(models.TransactionMillOperations).options(
        joinedload(models.TransactionMillOperations.transaction_allowance_deduction_details),
        joinedload(models.TransactionMillOperations.transaction_payments_mill_operations),
        joinedload(models.TransactionMillOperations.transaction_packaging_details).joinedload(models.TransactionPackagingDetails.packaging),
        joinedload(models.TransactionMillOperations.transaction_unloading_point_details),
        joinedload(models.TransactionMillOperations.transaction_stock_items).joinedload(models.TransactionStockItem.stock_items),
        joinedload(models.TransactionMillOperations.party),
        joinedload(models.TransactionMillOperations.broker),
        joinedload(models.TransactionMillOperations.transportor),
        joinedload(models.TransactionMillOperations.weight_bridge_operator),
    )

    # Filters
    if party_name:
        query = query.join(models.PartyDetails).filter(models.PartyDetails.party_name.ilike(f"%{party_name}%"))
    if broker_name:
        query = query.join(models.BrokerDetails).filter(models.BrokerDetails.broker_name.ilike(f"%{broker_name}%"))
    if transporter_name:
        query = query.join(models.TransportorDetails).filter(models.TransportorDetails.transportor_name.ilike(f"%{transporter_name}%"))
    if stock_item_name:
        query = query.join(models.TransactionStockItem).join(models.StockItems).filter(models.StockItems.stock_item_name.ilike(f"%{stock_item_name}%"))
    if start_date:
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        query = query.filter(models.TransactionMillOperations.transaction_date >= start)
    if end_date:
        end = datetime.strptime(end_date, "%Y-%m-%d").date()
        query = query.filter(models.TransactionMillOperations.transaction_date <= end)

    if transaction_type:
        query = query.filter(models.TransactionMillOperations.transaction_type == transaction_type)


    transactions = query.all()
    if not transactions:
        raise HTTPException(status_code=404, detail="No transactions found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Transaction Report"

    transaction_headers = [
        "Transaction ID", "Rst Number", "Bill Number", "Transaction Date", "Transaction Type",
        "Party Name", "Broker Name", "Transporter Name",
        "Gross Weight", "Tare Weight", "WB Operator", "Vehicle Number",
        "Net Total", "Payment Status", "Remaining Payment Amount"
    ]
    packaging_headers = ["Packaging Name", "Number of Bags", "Bag Weight (Grams)"]
    stock_item_headers = ["Item Name", "Bags", "Weight", "Rate"]
    payment_headers = ["Amount", "Date", "Narration"]
    allowance_headers = ["Amount", "Narration"]
    deduction_headers = ["Amount", "Narration"]

    # Row 1 & 2 headers
    ws.append(
        transaction_headers
        + packaging_headers
        + stock_item_headers
        + ["Payments", None, None]
        + ["Allowance", None]
        + ["Deduction", None]
        + ["Unloading Point", "Remarks"]
    )
    ws.append(
        [""] * len(transaction_headers)
        + packaging_headers
        + stock_item_headers
        + payment_headers
        + allowance_headers
        + deduction_headers
        + ["", ""]
    )

    # Merge header blocks
    for col_idx in range(1, len(transaction_headers) + 1):
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    pkg_start = len(transaction_headers) + 1
    pkg_end = pkg_start + len(packaging_headers) - 1
    ws.merge_cells(start_row=1, start_column=pkg_start, end_row=1, end_column=pkg_end)

    # ... (Rest of formatting logic)
    
    # Data Rows
    for txn in transactions:
        net_total, payment_status, remaining_payment = calculate_financials(
            txn.transaction_stock_items,
            txn.transaction_allowance_deduction_details,
            txn.transaction_payments_mill_operations
        )

        row_base = [
            txn.id, txn.rst_number, txn.bill_number, txn.transaction_date,
            "Purchase" if txn.transaction_type else "Sale",
            txn.party.party_name if txn.party else "",
            txn.broker.broker_name if txn.broker else "",
            txn.transportor.transportor_name if txn.transportor else "",
            txn.gross_weight, txn.tare_weight,
            txn.weight_bridge_operator.operator_name if txn.weight_bridge_operator else "",
            txn.vehicle_number,
            net_total, payment_status, remaining_payment
        ]

        # Determine max rows needed for this transaction (max of list lengths)
        n_pkg = len(txn.transaction_packaging_details)
        n_stock = len(txn.transaction_stock_items)
        n_pay = len(txn.transaction_payments_mill_operations)
        n_allow = len([x for x in txn.transaction_allowance_deduction_details if x.is_allowance])
        n_deduct = len([x for x in txn.transaction_allowance_deduction_details if not x.is_allowance])
        n_unload = len(txn.transaction_unloading_point_details)

        max_rows = max(n_pkg, n_stock, n_pay, n_allow, n_deduct, n_unload, 1)

        for i in range(max_rows):
            row = list(row_base) if i == 0 else [""] * len(row_base)

            # Packaging
            if i < n_pkg:
                p = txn.transaction_packaging_details[i]
                row.extend([p.packaging.packaging_name if p.packaging else "", p.bag_nos, p.packaging.bag_weight if p.packaging else ""])
            else:
                row.extend(["", "", ""])

            # Stock Items
            if i < n_stock:
                s = txn.transaction_stock_items[i]
                row.extend([s.stock_items.stock_item_name if s.stock_items else "", s.number_of_bags, s.weight, s.rate])
            else:
                row.extend(["", "", "", ""])

            # Payments
            if i < n_pay:
                pay = txn.transaction_payments_mill_operations[i]
                row.extend([pay.payment_amount, pay.payment_date, pay.payment_remarks])
            else:
                row.extend(["", "", ""])

            # Allowances
            allowances = [x for x in txn.transaction_allowance_deduction_details if x.is_allowance]
            if i < len(allowances):
                a = allowances[i]
                row.extend([a.allowance_deduction_amount, a.remarks])
            else:
                row.extend(["", ""])

            # Deductions
            deductions = [x for x in txn.transaction_allowance_deduction_details if not x.is_allowance]
            if i < len(deductions):
                d = deductions[i]
                row.extend([d.allowance_deduction_amount, d.remarks])
            else:
                row.extend(["", ""])
            
            # Unloading
            if i < n_unload:
                u = txn.transaction_unloading_point_details[i]
                row.extend([u.godown.godown_name if u.godown else "", u.remarks])
            else:
                row.extend(["", ""])

            ws.append(row)

    # Save to IO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=transaction_report.xlsx"}
    )
