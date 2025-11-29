import io
from fastapi import APIRouter, status, HTTPException, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import Optional, List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.db.session import get_db
from app.core.security import get_current_user
import app.models as models
import app.schemas.incoming_outgoing as schemas

router = APIRouter()

@router.get("/get_incoming_outgoing", response_model=List[schemas.IncomingOutgoingRead])
def get_incoming(
    brought_by: Optional[str] = None,
    vehicle_no: Optional[str] = None,
    party_through: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    query = db.query(models.IncomingOutgoing).options(
        joinedload(models.IncomingOutgoing.incoming_outgoing_items), joinedload(models.IncomingOutgoing.users)
    )

    if from_date:
        try:
            from_dt = datetime.strptime(from_date, "%Y-%m-%d").date()
            query = query.filter(models.IncomingOutgoing.io_date >= from_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid from_date format. Use YYYY-MM-DD.")

    if to_date:
        try:
            to_dt = datetime.strptime(to_date, "%Y-%m-%d").date()
            query = query.filter(models.IncomingOutgoing.io_date <= to_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid to_date format. Use YYYY-MM-DD.")
    
    if brought_by:
        query = query.filter(models.IncomingOutgoing.brought_by.ilike(f"%{brought_by}%"))

    if party_through:
        query = query.filter(models.IncomingOutgoing.party_through.ilike(f"%{party_through}%"))

    if vehicle_no:
        query = query.filter(models.IncomingOutgoing.vehicle_no.ilike(f"%{vehicle_no}%"))

    db_incoming = query.all()

    # Clean up None values in nested items
    for record in db_incoming:
        for item in record.incoming_outgoing_items:
            item.packaging = item.packaging if item.packaging is not None else 0
            item.weight_wb = item.weight_wb if item.weight_wb is not None else 0
            item.amount = item.amount if item.amount is not None else 0

        for payment in record.incoming_outgoing_payment or []:  # default to empty list
            payment.payment_amount = payment.payment_amount if payment.payment_amount is not None else 0
            payment.payment_date = payment.payment_date if payment.payment_date is not None else None

    return db_incoming

@router.get("/get_incoming_outgoing/{incoming_id}", response_model=schemas.IncomingOutgoingRead)
def get_incoming_by_id(
    incoming_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    db_incoming = db.query(models.IncomingOutgoing).options(
        joinedload(models.IncomingOutgoing.incoming_outgoing_items)
    ).filter(models.IncomingOutgoing.id == incoming_id).first()

    if not db_incoming:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Incoming or Outgoing not found")

    # Clean up None values in nested items
    for item in db_incoming.incoming_outgoing_items:
        item.packaging = item.packaging if item.packaging is not None else 0
        item.weight_wb = item.weight_wb if item.weight_wb is not None else 0
        item.amount = item.amount if item.amount is not None else 0

    for payment in db_incoming.incoming_outgoing_payment or []:  # default to empty list
        payment.payment_amount = payment.payment_amount if payment.payment_amount is not None else 0
        payment.payment_date = payment.payment_date if payment.payment_date is not None else None

    return db_incoming

@router.post("/", response_model=schemas.IncomingOutgoingRead)
def create_incoming(
    incoming: schemas.IncomingOutgoingCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # Extract non-nested fields
    incoming_data = incoming.model_dump(exclude={"incoming_outgoing_items", "incoming_outgoing_payment"})

    # Create parent record
    db_incoming = models.IncomingOutgoing(**incoming_data)
    db.add(db_incoming)
    db.commit()
    db.refresh(db_incoming)

    # Create nested items
    db_items = []
    for item in incoming.incoming_outgoing_items:
        db_item = models.IncomingOutgoingItems(
            incoming_outgoing_id=db_incoming.id,
            **item.model_dump()
        )
        db_items.append(db_item)
    db.add_all(db_items)

    # Create nested payments
    db_payments = []
    for payment in incoming.incoming_outgoing_payment:
        db_payment = models.IncomingOutgoingPayment(
            incoming_outgoing_id=db_incoming.id,
            payment_amount=payment.payment_amount if payment.payment_amount is not None else 0,
            payment_date=payment.payment_date if payment.payment_date is not None else None
        )
        db_payments.append(db_payment)
    db.add_all(db_payments)

    db.commit()
    db.refresh(db_incoming)

    return db_incoming

@router.put("/update_incoming/{incoming_id}", response_model=schemas.IncomingOutgoingRead)
def update_incoming(
    incoming_id: int,
    incoming: schemas.IncomingOutgoingCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    db_incoming = db.query(models.IncomingOutgoing).filter(models.IncomingOutgoing.id == incoming_id).first()
    if not db_incoming:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No Incoming or Outgoing found")

    # Update parent fields (excluding nested lists)
    incoming_data = incoming.model_dump(exclude={"incoming_outgoing_items", "incoming_outgoing_payment"}, exclude_unset=True)
    for key, value in incoming_data.items():
        setattr(db_incoming, key, value)

    # Update nested items
    if incoming.incoming_outgoing_items:
        # Delete existing items
        db.query(models.IncomingOutgoingItems).filter(models.IncomingOutgoingItems.incoming_outgoing_id == db_incoming.id).delete()
        # Add new items
        for item in incoming.incoming_outgoing_items:
            db_item = models.IncomingOutgoingItems(
                incoming_outgoing_id=db_incoming.id,
                **item.model_dump()
            )
            db.add(db_item)

    # Update nested payments
    if incoming.incoming_outgoing_payment:
        db.query(models.IncomingOutgoingPayment).filter(models.IncomingOutgoingPayment.incoming_outgoing_id == db_incoming.id).delete()
        for payment in incoming.incoming_outgoing_payment:
            db_payment = models.IncomingOutgoingPayment(
                incoming_outgoing_id=db_incoming.id,
                **payment.model_dump()
            )
            db.add(db_payment)

    db.commit()
    db.refresh(db_incoming)

    return db_incoming

@router.put("/update_outgoing/{outgoing_id}", response_model=schemas.IncomingOutgoingRead)
def update_outgoing(
    outgoing_id: int,
    outgoing: schemas.IncomingOutgoingCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # 1. Get the existing record
    db_outgoing = db.query(models.IncomingOutgoing).filter(
        models.IncomingOutgoing.id == outgoing_id
    ).first()
    if not db_outgoing:
        raise HTTPException(status_code=404, detail="No Incoming or Outgoing found")

    # 2. Update main fields (excluding nested lists)
    outgoing_data = outgoing.model_dump(exclude={"incoming_outgoing_items", "incoming_outgoing_payment"}, exclude_unset=True)
    for key, value in outgoing_data.items():
        setattr(db_outgoing, key, value)
    db.add(db_outgoing)
    db.commit()
    db.refresh(db_outgoing)

    # 3. Update items
    if outgoing.incoming_outgoing_items is not None:
        # Delete existing items
        db.query(models.IncomingOutgoingItems).filter(
            models.IncomingOutgoingItems.incoming_outgoing_id == db_outgoing.id
        ).delete()
        db.commit()

        # Add new items
        db_items = []
        for item in outgoing.incoming_outgoing_items:
            db_item = models.IncomingOutgoingItems(
                incoming_outgoing_id=db_outgoing.id,
                **item.model_dump()
            )
            db.add(db_item)
            db_items.append(db_item)
        db.commit()
        for i in db_items:
            db.refresh(i)
    else:
        db_items = db_outgoing.incoming_outgoing_items

    # 4. Update payments (optional, same pattern)
    if outgoing.incoming_outgoing_payment is not None:
        # Delete existing payments
        db.query(models.IncomingOutgoingPayment).filter(
            models.IncomingOutgoingPayment.incoming_outgoing_id == db_outgoing.id
        ).delete()
        db.commit()

        # Add new payments
        db_payments = []
        for pay in outgoing.incoming_outgoing_payment:
            db_payment = models.IncomingOutgoingPayment(
                incoming_outgoing_id=db_outgoing.id,
                payment_amount=pay.payment_amount,
                payment_date=pay.payment_date
            )
            db.add(db_payment)
            db_payments.append(db_payment)
        db.commit()
        for p in db_payments:
            db.refresh(p)
    else:
        db_payments = db_outgoing.incoming_outgoing_payment

    # 5. Construct Pydantic response
    response = schemas.IncomingOutgoingRead(
        id=db_outgoing.id,
        io_date=db_outgoing.io_date,
        is_incoming=db_outgoing.is_incoming,
        rst_bill=db_outgoing.rst_bill,
        brought_by=db_outgoing.brought_by,
        mob_no=db_outgoing.mob_no,
        vehicle_no=db_outgoing.vehicle_no,
        origin=db_outgoing.origin,
        party_through=db_outgoing.party_through,
        transportation_expense=db_outgoing.transportation_expense,
        remarks=db_outgoing.remarks,
        user_login_id=db_outgoing.user_login_id,
        time_stamp=db_outgoing.time_stamp,
        users=db_outgoing.users,  # ensure compatible with UserResponse
        incoming_outgoing_items=[
            schemas.IncomingOutgoingItem(
                id=i.id,
                incoming_outgoing_id=i.incoming_outgoing_id,
                jins=i.jins,
                bags_no=i.bags_no,
                quantity=i.quantity,
                packaging=i.packaging,
                weight_society=i.weight_society,
                weight_wb=i.weight_wb,
                amount=i.amount
            ) for i in db_items
        ],
        incoming_outgoing_payment=[
            schemas.IncomingOutgoingPayments(
                payment_amount=p.payment_amount,
                payment_date=p.payment_date
            ) for p in db_payments
        ]
    )

    return response

@router.get("/get_all_incomings", response_model=List[schemas.IncomingOutgoingRead])
def get_all_incomings(
    brought_by: Optional[str] = None,
    party_through: Optional[str] = None,
    vehicle_no: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    query = db.query(models.IncomingOutgoing).options(
        joinedload(models.IncomingOutgoing.incoming_outgoing_items)
    ).filter(models.IncomingOutgoing.is_incoming == True)

    if from_date:
        try:
            from_dt = datetime.strptime(from_date, "%Y-%m-%d").date()
            query = query.filter(models.IncomingOutgoing.io_date >= from_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid from_date format. Use YYYY-MM-DD.")

    if to_date:
        try:
            to_dt = datetime.strptime(to_date, "%Y-%m-%d").date()
            query = query.filter(models.IncomingOutgoing.io_date <= to_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid to_date format. Use YYYY-MM-DD.")
        
    if brought_by:
        query = query.filter(models.IncomingOutgoing.brought_by.ilike(f"%{brought_by}%"))

    if party_through:
        query = query.filter(models.IncomingOutgoing.party_through.ilike(f"%{party_through}%"))

    if vehicle_no:
        query = query.filter(models.IncomingOutgoing.vehicle_no.ilike(f"%{vehicle_no}%"))

    results = query.all()

    return results

@router.get("/download_incoming_report")
def download_incoming_report(
    brought_by: Optional[str] = None,
    party_through: Optional[str] = None,
    vehicle_no: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    query = db.query(models.IncomingOutgoing).options(
        joinedload(models.IncomingOutgoing.incoming_outgoing_items)
    ).filter(models.IncomingOutgoing.is_incoming == True)

    if from_date:
        try:
            from_dt = datetime.strptime(from_date, "%Y-%m-%d").date()
            query = query.filter(models.IncomingOutgoing.io_date >= from_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid from_date format. Use YYYY-MM-DD.")

    if to_date:
        try:
            to_dt = datetime.strptime(to_date, "%Y-%m-%d").date()
            query = query.filter(models.IncomingOutgoing.io_date <= to_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid to_date format. Use YYYY-MM-DD.")
        
    if brought_by:
        query = query.filter(models.IncomingOutgoing.brought_by.ilike(f"%{brought_by}%"))

    if party_through:
        query = query.filter(models.IncomingOutgoing.party_through.ilike(f"%{party_through}%"))

    if vehicle_no:
        query = query.filter(models.IncomingOutgoing.vehicle_no.ilike(f"%{vehicle_no}%"))

    results = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Incoming Report"

    main_headers = [
        "ID", "TID", "I/O DATE", "RST BILL NO.", "BROUGHT BY", "MOBILE NO.", "VEHICLE NO.",
        "ORIGIN", "PARTY THROUGH", "TRANSPORT EXP. ₹", "PAYMENT ₹", "PAYMENT DATE",
        "ENTRY BY"
    ]
    item_headers = ["JINS", "BAGS/NOS", "QUANTITY", "PACKAGING", "WEIGHT (SOCIETY)", "WEIGHT (BRIDGE)", "AMOUNT ₹"]
    total_columns = len(main_headers) + len(item_headers)

    # Header Rows
    ws.append([""] * total_columns)
    ws.append([""] * total_columns)

    # Merge and write headers
    for col_idx, header in enumerate(main_headers):
        ws.merge_cells(start_row=1, start_column=col_idx + 1, end_row=2, end_column=col_idx + 1)
        ws.cell(row=1, column=col_idx + 1).value = header

    ws.merge_cells(start_row=1, start_column=len(main_headers) + 1, end_row=1, end_column=total_columns)
    ws.cell(row=1, column=len(main_headers) + 1).value = "ITEMS"

    for i, header in enumerate(item_headers):
        ws.cell(row=2, column=len(main_headers) + i + 1).value = header

    # Style headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=total_columns):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

    # Fill data
    data_start_row = 3
    row_number = 1  # Global index for ID column
    current_row = data_start_row
    for entry in results:
        payments = entry.incoming_outgoing_payment or []
        items = entry.incoming_outgoing_items or [None]
        item_count = len(items)

        for idx, item in enumerate(items):
            row = []
            if idx == 0:
                row.extend([
                    row_number,
                    entry.id,
                    str(entry.io_date),
                    entry.rst_bill or "",
                    entry.brought_by or "",
                    entry.mob_no or "",
                    entry.vehicle_no or "",
                    entry.origin or "",
                    entry.party_through or "",
                    entry.transportation_expense or "",
                    payments[0].payment_amount if payments else "",
                    str(payments[0].payment_date) if payments and payments[0].payment_date else "",
                    entry.user_login_id or ""
                ])
            else:
                row.extend([""] * len(main_headers))

            row.extend([
                item.jins if item else "",
                "BAGS" if item and item.bags_no else "NOS", 
                item.quantity if item else "",
                item.packaging if item else "",
                item.weight_society if item else "",
                item.weight_wb if item else "",
                item.amount if item else ""
            ])
            ws.append(row)

        # Merge main columns for entries with multiple items
        if item_count > 1:
            for col_idx in range(1, len(main_headers)+1):
                ws.merge_cells(
                    start_row=current_row,
                    end_row=current_row + item_count - 1,
                    start_column=col_idx,
                    end_column=col_idx
                )
        current_row += item_count
        row_number += 1  # Increment global ID per transaction

    # Style data rows with alignment and currency formatting
    row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    currency_format = '₹#,##,##0.00'  # Indian rupee format with lakh/crore grouping

    # Columns that should be right-aligned (numeric/currency columns)
    right_align_columns = [
        10,  # Transport Exp. ₹
        11,  # Payment ₹
        len(main_headers) + 3,  # Quantity
        len(main_headers) + 5,  # Weight (Society)
        len(main_headers) + 6,  # Weight (Bridge)
        len(main_headers) + 7,  # Amount ₹
    ]

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=total_columns):
        for col_idx, cell in enumerate(row, start=1):
            # Fill color
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Apply right/left alignment
            if col_idx in right_align_columns:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Apply currency format where relevant
            if col_idx in [10, 11, len(main_headers) + 7]:
                try:
                    if isinstance(cell.value, (int, float)) and cell.value != "-":
                        cell.number_format = currency_format
                except Exception:
                    pass

    # Add borders
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=total_columns):
        for cell in row:
            cell.border = border

    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    # Return Excel
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=incoming_report.xlsx"}
    )

@router.get("/download_outgoing_report")
def download_outgoing_report(
    brought_by: Optional[str] = None,
    party_through: Optional[str] = None,
    vehicle_no: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    query = db.query(models.IncomingOutgoing).options(
        joinedload(models.IncomingOutgoing.incoming_outgoing_items)
    ).filter(models.IncomingOutgoing.is_incoming == False)

    if from_date:
        try:
            from_dt = datetime.strptime(from_date, "%Y-%m-%d").date()
            query = query.filter(models.IncomingOutgoing.io_date >= from_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid from_date format. Use YYYY-MM-DD.")

    if to_date:
        try:
            to_dt = datetime.strptime(to_date, "%Y-%m-%d").date()
            query = query.filter(models.IncomingOutgoing.io_date <= to_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid to_date format. Use YYYY-MM-DD.")
        
    if brought_by:
        query = query.filter(models.IncomingOutgoing.brought_by.ilike(f"%{brought_by}%"))

    if party_through:
        query = query.filter(models.IncomingOutgoing.party_through.ilike(f"%{party_through}%"))

    if vehicle_no:
        query = query.filter(models.IncomingOutgoing.vehicle_no.ilike(f"%{vehicle_no}%"))

    results = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Outgoing Report"

    main_headers = [
        "ID", "TID", "I/O DATE", "RST BILL NO.", "BROUGHT BY", "MOBILE NO.", "VEHICLE NO.",
        "ORIGIN", "PARTY THROUGH", "TRANSPORT EXP. ₹", "PAYMENT ₹", "PAYMENT DATE",
        "ENTRY BY"
    ]
    item_headers = ["JINS", "BAGS/NOS", "QUANTITY", "PACKAGING", "WEIGHT (SOCIETY)", "WEIGHT (BRIDGE)", "AMOUNT ₹"]
    total_columns = len(main_headers) + len(item_headers)

    # Header Rows
    ws.append([""] * total_columns)
    ws.append([""] * total_columns)

    # Merge and write headers
    for col_idx, header in enumerate(main_headers):
        ws.merge_cells(start_row=1, start_column=col_idx + 1, end_row=2, end_column=col_idx + 1)
        ws.cell(row=1, column=col_idx + 1).value = header

    ws.merge_cells(start_row=1, start_column=len(main_headers) + 1, end_row=1, end_column=total_columns)
    ws.cell(row=1, column=len(main_headers) + 1).value = "ITEMS"

    for i, header in enumerate(item_headers):
        ws.cell(row=2, column=len(main_headers) + i + 1).value = header

    # Style headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=total_columns):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

    # Fill data
    data_start_row = 3
    row_number = 1  # global index for ID column
    current_row = data_start_row
    for entry in results:
        items = entry.incoming_outgoing_items or [None]
        item_count = len(items)

        for idx, item in enumerate(items):
            row = []
            if idx == 0:
                row.extend([
                    row_number,
                    entry.id,
                    str(entry.io_date),
                    entry.rst_bill or "",
                    entry.brought_by or "",
                    entry.mob_no or "",
                    entry.vehicle_no or "",
                    entry.origin or "",
                    entry.party_through or "",
                    entry.transportation_expense or "",
                    entry.payment or "",
                    str(entry.payment_date) if entry.payment_date else "",
                    entry.user_login_id or ""
                ])
            else:
                row.extend([""] * len(main_headers))

            row.extend([
                item.jins if item else "",
                "BAGS" if item and item.bags_no else "NOS", 
                item.quantity if item else "",
                item.packaging if item else "",
                item.weight_society if item else "",
                item.weight_wb if item else "",
                item.amount if item else ""
            ])
            ws.append(row)

        # Merge main columns for entries with multiple items
        if item_count > 1:
            for col_idx in range(1, len(main_headers) + 1):
                ws.merge_cells(
                    start_row=current_row,
                    end_row=current_row + item_count - 1,
                    start_column=col_idx,
                    end_column=col_idx
                )
        current_row += item_count
        row_number += 1  # increment ID per transaction

    # White fill + left align for data rows
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=total_columns):
        for cell in row:
            cell.fill = white_fill
            cell.alignment = left_align

    # Borders
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=total_columns):
        for cell in row:
            cell.border = border

    # Auto-fit
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    # Return Excel
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=outgoing_report.xlsx"}
    )
